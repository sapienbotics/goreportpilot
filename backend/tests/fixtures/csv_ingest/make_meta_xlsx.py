"""
Build meta_ads_multisheet.xlsx.

Two sheets. The first is a decoy: a small key/value config block that a naive
"take the first sheet" importer would pick up instead of the real data. The
second holds the campaign table, with CTR stored the way Meta actually exports
it — as a decimal fraction (0.0047), not a percentage.

    cd backend && python tests/fixtures/csv_ingest/make_meta_xlsx.py
"""
import os
from datetime import date

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "meta_ads_multisheet.xlsx")

# (reporting_start, campaign, spend_inr, impressions, link_clicks, ctr_fraction, results)
ROWS = [
    (date(2026, 7, 1),  "IN | Prospecting | Broad",      47512.40, 1_284_500, 6037, 0.0047, 312),
    (date(2026, 7, 2),  "IN | Prospecting | Broad",      51203.75, 1_402_180, 7291, 0.0052, 358),
    (date(2026, 7, 3),  "IN | Retargeting | 30d",        44890.20, 1_105_940, 4866, 0.0044, 401),
    (date(2026, 7, 4),  "IN | Retargeting | 30d",        49876.50, 1_233_760, 6045, 0.0049, 427),
    (date(2026, 7, 5),  "IN | Lookalike | 1pct",         46320.15, 1_051_320, 5362, 0.0051, 289),
    (date(2026, 7, 6),  "IN | Lookalike | 1pct",         47100.00, 1_180_400, 5430, 0.0046, 305),
    (date(2026, 7, 7),  "IN | Prospecting | Broad",      45989.00, 1_002_640, 5314, 0.0053, 333),
]

SETTINGS = [
    ("Account ID", "act_1094857362"),
    ("Currency", "INR"),
    ("Time zone", "Asia/Kolkata"),
    ("Attribution setting", "7-day click, 1-day view"),
]


def main() -> None:
    workbook = openpyxl.Workbook()

    decoy = workbook.active
    decoy.title = "Account Settings"
    decoy.append(["Setting", "Value"])
    for key, value in SETTINGS:
        decoy.append([key, value])

    sheet = workbook.create_sheet("Campaign Performance")
    sheet.append([
        "Reporting starts", "Campaign name", "Amount spent (INR)",
        "Impressions", "Link clicks", "CTR (all)", "Results",
    ])
    for row in ROWS:
        sheet.append(list(row))

    workbook.save(OUT)

    total_spend = sum(r[2] for r in ROWS)
    mean_ctr = sum(r[5] for r in ROWS) / len(ROWS)
    print(f"wrote {OUT}")
    print(f"  sheets:        {workbook.sheetnames}")
    print(f"  total spend:   {total_spend:,.2f} INR")
    print(f"  mean CTR:      {mean_ctr:.6f} (= {mean_ctr * 100:.4f}%)")


if __name__ == "__main__":
    main()
